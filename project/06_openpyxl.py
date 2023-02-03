import openpyxl
from openpyxl import Workbook


#1 openpy1
    # wb = Workbook() 
    # ws = wb.active
    # ws.title = 'openpy'
    # wb.save('openpy01.xlsx')
    # wb.close

#2 openpy2 - 새로운 시트 생성
    # wb = Workbook() 
    # ws2 = wb.create_sheet()
    # ws2.title = 'openpy2'

    # ws3 = wb.create_sheet('openpy3')

    # wb['openpy2'].title = 'new openpy2'     # 제목에 대한 것을 Key로 사용가능
    # ws2 = wb['new openpy2']
    # ws2['A3'] = 'coke'  # 셀에 대한 것을 Key로 사용가능 =>  셀에 대한 조작

    # wb.save('openpy02.xlsx')

#3 openpy3 - 기존에 있던 워크북을 가져오기 & 셀 데이터 입력
    # wb = openpyxl.load_workbook(filename='openpy.xlsx')
    # ws1 = wb.active

    # # 셀 데이터 접근
    #     # 1. Key로 접근 : 시트명['주소']   -    행(row)- A,B,C,D,... / 열(column) - 1,2,3,4,5,6,...
    # for i in range(1,10):
    #     ws1[f'A{i}'] = 'coke'
    #     ws1[f'b{i}'] = 'pizza'
    #     ws1[f'C{i}'] = 'chicken'
        
    # print(ws1['b5']) # 셀 객체 자체가 출력 => value 를 통해 셀의 값을 출력해야함
    # print(ws1['b5'].value)
    # print(ws1['f5'].value) # 없는 값에 대해서는 None이 출력

    #     #2. cell함수로 접근 : 시트명.cell(x,y)       - 행(row) - 1,2,3,4,5,.... / 열(column) - 1,2,3,4,5,6,...
    # print(ws1.celll(1,1).value)

    # wb.save('openpy.xlsx')

#4. practice - 구구단 만들기
    # wb = openpyxl.load_workbook(filename = 'openpy.xlsx')

    # ws2 = wb.create_sheet('gugu1')
    # for i in range(1,10):
    #     for j in range(1,10):
    #         ws2.cell(i,j).value = i*j
            
    # ws3 = wb.create_sheet('gugu2')
    # for i in range(1,10):
    #     for j in range(1,10):
    #         ws3[f'{chr(i+64)}{j}'] = i*j

    # wb.save('openpy.xlsx')
    
#5. openpy4 - 여러 개의 범위 데이터를 불러오기
    # wb = openpyxl.load_workbook(filename = './resource/sample.xlsx')
    # ws = wb.active

    #    #5-1. 모든 행, 모든 열의 값 불러오기
    # print(ws.max_row) # 시트 내 최대 행 위치
    # print(ws.max_column)    # 시트 내 최대 열 위치
    # print(ws.min_row) # 시트 내 시작 행 위치    => 첫 시작이 1이 아닐 수도 있으므로
    # print(ws.min_column) # 시트 내 시작 열 위치

    # for i in range(1,ws.max_row+1):
    #     for j in range(1,ws.max_column+1):
    #         print(ws.cell(i,j).value, end = ' ')
    #     print('')
    
#5-2. range - 기본적으로 column을 통해 슬라이싱이 수행됨(열별로 튜플 형식으로 출력)    Ex)년(A열)과 월(B열)만 가져오기

    #     # (1) 열 단위로 불러오기
    #     col_range = ws['A:B']
    #     for cols in col_range:
    #         for cell in cols:
    #             print(cell.value)
    #         print("---")

    #     # (2) 행 단위로 불러오기
    #     row_range = ws[1:4]
    #     for rows in row_range:
    #         for cell in rows:
    #             print(cell.value)
    #         print('---')

    # # (3) 지정 범위로 불러오기 => 주로 사용
    # row_col_range = ws['A1:B4']

    # for cols in row_col_range:
    #     for cell in cols:
    #         print(cell.value, end = ' ')
    #     print()

    # # print([i for i in zip(col_range)])


# 6. openpy 5 - iter_rows(행을 기준으로 반복-> 한행씩 가져오기)와 iter_columns를 이용한 반복문 => 조건식 등에 대한 연산에 용이
    # wb = openpyxl.load_workbook(filename = './resource/sample.xlsx')
    # ws = wb.active
    # for row in ws.iter_rows(min_row = 1, max_row = 4,min_col =1,max_col =4):
    #     print(row[0].value,row[1].value,row[2].value)
    

# 7. practice - 생산량이 3,500,000 이상인 년도와 월을 출력하려면?
    # wb = openpyxl.load_workbook(filename = './resource/sample.xlsx')
    # ws = wb.active
    #     #7-1
    # for row in ws.iter_rows(min_row=2,max_row = ws.max_row):
    #     if int(row[2].value) >=3500000 :
    #         print(row[0].value, row[1].value)
    # print()

    #     # 7-2
    # for row in ws[2:ws.max_row]:
    #     if int(row[2].value) >= 3500000:
    #         print(row[0].value, row[1].value)